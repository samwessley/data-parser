import pandas as pd
import openpyxl
import os
from openpyxl.styles import PatternFill, Font, Alignment, NamedStyle
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import sys
import logging

# Set up logging for debugging
logging.basicConfig(level=logging.INFO)

def find_last_row(input_file_path, sheet_name):
    """Find the last consecutive non-empty row in column A with data, starting from the top."""
    workbook = openpyxl.load_workbook(input_file_path, data_only=True)
    sheet = workbook[sheet_name]
    
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value is None:
            return row - 1
    return sheet.max_row

def read_excel(input_file_path, sheet_name, last_row):
    df = pd.read_excel(
        input_file_path, 
        sheet_name=sheet_name, 
        usecols="A:G", 
        nrows=last_row,
        engine='openpyxl'
    )
    return df

def read_full_sheet(input_file_path, sheet_name):
    df = pd.read_excel(
        input_file_path,
        sheet_name=sheet_name,
        engine='openpyxl'
    )
    return df

def write_excel(df1, df2, df3, output_file_path):
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        pd.DataFrame().to_excel(writer, sheet_name='Instructions', index=False)
        pd.DataFrame().to_excel(writer, sheet_name='Data Validation Tests', index=False)
        df1.to_excel(writer, sheet_name='Comparative Trial Balances', index=False)
        df2.to_excel(writer, sheet_name='Journal Entries & Lines', index=False)
        df3.to_excel(writer, sheet_name='Mapping Categories', index=False)

    workbook = openpyxl.load_workbook(output_file_path)
    format_header_and_second_row(workbook, "Comparative Trial Balances")
    format_header_and_second_row_jel(workbook, "Journal Entries & Lines")
    apply_accounting_format(workbook, "Comparative Trial Balances", [3, 4])
    apply_accounting_and_date_format_jel(workbook, "Journal Entries & Lines")
    adjust_column_widths(workbook, "Comparative Trial Balances")
    adjust_column_widths(workbook, "Journal Entries & Lines")
    workbook.save(output_file_path)

def format_header_and_second_row(workbook, sheet_name):
    sheet = workbook[sheet_name]
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    second_row_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center")

    for col in range(1, 8):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for col in range(1, 8):
        cell = sheet.cell(row=2, column=col)
        cell.fill = second_row_fill
        cell.font = header_font
        cell.alignment = center_alignment

def format_header_and_second_row_jel(workbook, sheet_name):
    sheet = workbook[sheet_name]
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    light_blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    grey_fill = PatternFill(start_color="999999", end_color="999999", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_alignment = Alignment(horizontal="center")

    for col in range(1, 8):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for col in [1, 3, 4, 6, 7]:
        cell = sheet.cell(row=2, column=col)
        cell.fill = light_blue_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for col in [2, 5]:
        cell = sheet.cell(row=2, column=col)
        cell.fill = grey_fill
        cell.font = header_font
        cell.alignment = center_alignment

def apply_accounting_format(workbook, sheet_name, columns):
    sheet = workbook[sheet_name]
    accounting_style = NamedStyle(name="accounting_style", number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')

    if "accounting_style" not in workbook.named_styles:
        workbook.add_named_style(accounting_style)
    else:
        logging.info("Accounting style already exists.")

    for col in columns:
        for row in range(3, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.style = "accounting_style"

def apply_accounting_and_date_format_jel(workbook, sheet_name):
    sheet = workbook[sheet_name]
    accounting_style = NamedStyle(name="accounting_style", number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')
    date_style = NamedStyle(name="date_style", number_format='m/d/yy')

    if "accounting_style" not in workbook.named_styles:
        workbook.add_named_style(accounting_style)
    else:
        logging.info("Accounting style already exists in 'Journal Entries & Lines'.")

    if "date_style" not in workbook.named_styles:
        workbook.add_named_style(date_style)
    else:
        logging.info("Date style already exists.")

    for col in [6, 7]:
        for row in range(3, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=col)
            cell.style = "accounting_style"

    for row in range(3, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=3)
        cell.style = "date_style"

def adjust_column_widths(workbook, sheet_name):
    sheet = workbook[sheet_name]
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

def validate_comparative_trial_balances(ctb_df, mapping_df, output_file_path):
    valid_categories = set(mapping_df.iloc[0:, 0].dropna().astype(str).str.strip())
    workbook = openpyxl.load_workbook(output_file_path)
    sheet = workbook["Comparative Trial Balances"]
    highlight_fill_e = PatternFill(start_color="F7B4AE", end_color="F7B4AE", fill_type="solid")
    highlight_fill_f = PatternFill(start_color="F7B4AE", end_color="F7B4AE", fill_type="solid")

    invalid_entries = []
    for index, (category, value) in enumerate(zip(ctb_df.iloc[2:, 4].dropna().astype(str).str.strip(),
                                                  ctb_df.iloc[2:, 5].dropna().astype(str).str.strip()), start=3):
        if category not in valid_categories:
            invalid_entries.append((index + 1, category))
            sheet.cell(row=index + 1, column=5).fill = highlight_fill_e
            sheet.cell(row=index + 1, column=6).fill = highlight_fill_f
        else:
            column_mapping = {
                "Assets": 1,
                "Liabilities": 2,
                "Equity": 3,
                "Income": 4,
                "Expenses": 5
            }
            mapping_column = column_mapping.get(category)
            if mapping_column is not None:
                valid_values = set(mapping_df.iloc[0:, mapping_column].dropna().astype(str).str.strip())
                if value not in valid_values:
                    invalid_entries.append((index + 1, category, value))
                    sheet.cell(row=index + 1, column=6).fill = highlight_fill_f

    workbook.save(output_file_path)
    return invalid_entries

def check_balance_sums(ctb_df):
    sum_c = round(ctb_df.iloc[1:, 2].sum(), 2)
    sum_d = round(ctb_df.iloc[1:, 3].sum(), 2)
    if sum_c != 0:
        return "Prior Period Balance does not sum to 0."
    if sum_d != 0:
        return "Current Period Balance does not sum to 0."
    return "Prior Period Balance and Current Period Balance both sum to 0. ✅"

def process_journal_entries(workbook, sheet_name):
    sheet = workbook[sheet_name]
    last_row = sheet.max_row

    for row in range(3, last_row + 1):
        value_f = sheet.cell(row=row, column=6).value or 0
        value_g = sheet.cell(row=row, column=7).value or 0
        net_value = value_f - value_g

        if net_value > 0:
            sheet.cell(row=row, column=6).value = net_value
            sheet.cell(row=row, column=7).value = 0
        elif net_value < 0:
            sheet.cell(row=row, column=6).value = 0
            sheet.cell(row=row, column=7).value = -net_value
        else:
            sheet.cell(row=row, column=6).value = 0
            sheet.cell(row=row, column=7).value = 0

def check_debit_credit_sums(workbook, sheet_name):
    sheet = workbook[sheet_name]
    last_row = sheet.max_row
    sum_f = sum(sheet.cell(row=row, column=6).value or 0 for row in range(3, last_row + 1))
    sum_g = sum(sheet.cell(row=row, column=7).value or 0 for row in range(3, last_row + 1))

    sum_f_str = f"${sum_f:,.2f}"
    sum_g_str = f"${sum_g:,.2f}"

    if round(sum_f, 2) == round(sum_g, 2):
        return f"The sums of the Debit and Credit columns are equal. Both are {sum_f_str}. ✅"
    else:
        return f"The sums of the Debit and Credit columns are not equal. The Debit column sums to {sum_f_str} while the Credit column sums to {sum_g_str}. ❌"

def check_journal_entry_balances(workbook, sheet_name, output_directory):
    sheet = workbook[sheet_name]
    last_row = sheet.max_row

    journal_entries = {}
    for row in range(3, last_row + 1):
        journal_id = sheet.cell(row=row, column=1).value
        date = sheet.cell(row=row, column=3).value
        debit = sheet.cell(row=row, column=6).value or 0
        credit = sheet.cell(row=row, column=7).value or 0

        if journal_id not in journal_entries:
            journal_entries[journal_id] = {"debit": 0, "credit": 0, "dates": set()}

        journal_entries[journal_id]["debit"] += debit
        journal_entries[journal_id]["credit"] += credit
        journal_entries[journal_id]["dates"].add(date)

    je_list_path = os.path.join(output_directory, 'je_list.xlsx')
    je_workbook = openpyxl.Workbook()
    je_sheet = je_workbook.active
    je_sheet.title = "Journal Entries"

    headers = ["Journal ID", "Net Balance", "Earliest Date", "Latest Date", "Date Difference (Days)"]
    for col_num, header in enumerate(headers, start=1):
        je_sheet.cell(row=1, column=col_num, value=header)

    all_balanced = True
    for row_num, (journal_id, data) in enumerate(journal_entries.items(), start=2):
        net_balance = round(data["debit"] - data["credit"], 2)
        earliest_date = min(data["dates"])
        latest_date = max(data["dates"])
        date_difference = (latest_date - earliest_date).days

        if net_balance != 0 or len(data["dates"]) > 1:
            all_balanced = False

        je_sheet.cell(row=row_num, column=1, value=journal_id)
        je_sheet.cell(row=row_num, column=2, value=net_balance)
        je_sheet.cell(row=row_num, column=3, value=earliest_date)
        je_sheet.cell(row=row_num, column=4, value=latest_date)
        je_sheet.cell(row=row_num, column=5, value=date_difference)

    accounting_style = NamedStyle(name="accounting_style_je", number_format='_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')
    if "accounting_style_je" not in je_workbook.named_styles:
        je_workbook.add_named_style(accounting_style)

    for row in range(2, je_sheet.max_row + 1):
        je_sheet.cell(row=row, column=2).style = "accounting_style_je"

    adjust_column_widths(je_workbook, "Journal Entries")
    je_workbook.save(je_list_path)

    if all_balanced:
        return "All journal entries balance. ✅"
    else:
        return "There are unbalanced journal entries. See 'je_list.xlsx' ❌"

def check_account_id_in_ctb(workbook, jel_df, ctb_df, output_file_path):
    jel_account_ids = jel_df['Account ID'].astype(str).unique()
    ctb_account_ids = ctb_df['Account ID'].astype(str).tolist()

    workbook_ctb = openpyxl.load_workbook(output_file_path)
    sheet_ctb = workbook_ctb["Comparative Trial Balances"]
    highlight_fill = PatternFill(start_color="F7B4AE", end_color="F7B4AE", fill_type="solid")

    for account_id in jel_account_ids:
        if account_id not in ctb_account_ids:
            sheet_ctb.cell(row=sheet_ctb.max_row + 1, column=1, value=account_id)
            sheet_ctb.cell(row=sheet_ctb.max_row, column=2, value=account_id)
            zero_cell_c = sheet_ctb.cell(row=sheet_ctb.max_row, column=3, value=0)
            zero_cell_c.style = "accounting_style"
            zero_cell_d = sheet_ctb.cell(row=sheet_ctb.max_row, column=4, value=0)
            zero_cell_d.style = "accounting_style"

            for col in range(5, 8):
                sheet_ctb.cell(row=sheet_ctb.max_row, column=col).fill = highlight_fill

            ctb_account_ids.append(account_id)

    workbook_ctb.save(output_file_path)

def select_file():
    global input_file_path
    input_file_path = filedialog.askopenfilename(
        title="Select the source Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if input_file_path:
        log_output.insert(tk.END, f"Selected file: {input_file_path}\n")
        validate_button.config(state=tk.NORMAL)

def run_validation():
    if not input_file_path:
        messagebox.showerror("Error", "Please select a file first.")
        return

    output_file_name = 'output_file.xlsx'
    output_file_path = os.path.join(os.path.expanduser('~'), 'Downloads', output_file_name)

    log_output.insert(tk.END, "----COMPARATIVE TRIAL BALANCE TAB----\n")

    last_row_ctb = find_last_row(input_file_path, "Comparative Trial Balances")
    df_ctb = read_excel(input_file_path, "Comparative Trial Balances", last_row_ctb)
    last_row_jel = find_last_row(input_file_path, "Journal Entries & Lines")
    df_jel = read_excel(input_file_path, "Journal Entries & Lines", last_row_jel)
    df_mapping = read_full_sheet(input_file_path, "Mapping Categories")

    write_excel(df_ctb, df_jel, df_mapping, output_file_path)
    workbook = openpyxl.load_workbook(output_file_path)

    process_journal_entries(workbook, "Journal Entries & Lines")
    workbook.save(output_file_path)

    invalid_entries = validate_comparative_trial_balances(df_ctb, df_mapping, output_file_path)
    balance_check = check_balance_sums(df_ctb)
    log_output.insert(tk.END, balance_check + "\n")

    if invalid_entries:
        log_output.insert(tk.END, "Invalid mappings found in the 'Comparative Trial Balances' tab:\n")
        for entry in invalid_entries:
            if len(entry) == 2:
                row, category = entry
                log_output.insert(tk.END, f"Row {row}: Category '{category}' not found in 'Mapping Categories'\n")
            else:
                row, category, value = entry
                log_output.insert(tk.END, f"Row {row}: Value '{value}' not valid for category '{category}' in 'Mapping Categories'\n")
    else:
        log_output.insert(tk.END, "All Account Types match one of the options on the Mapping Categories tab. ✅\n")
        log_output.insert(tk.END, "All Account Mappings match one of the options on the Mapping Categories tab. ✅\n")
        log_output.insert(tk.END, "All Account Mappings have a matching Account Type. ✅\n")

    log_output.insert(tk.END, "----JOURNAL ENTRIES & LINES TAB----\n")
    debit_credit_check = check_debit_credit_sums(workbook, "Journal Entries & Lines")
    log_output.insert(tk.END, debit_credit_check + "\n")
    journal_entry_check = check_journal_entry_balances(workbook, "Journal Entries & Lines", os.path.expanduser('~'))
    log_output.insert(tk.END, journal_entry_check + "\n")
    check_account_id_in_ctb(workbook, df_jel, df_ctb, output_file_path)
    download_button.config(state=tk.NORMAL)

def download_file():
    save_path = filedialog.asksaveasfilename(
        initialdir=os.path.expanduser('~'),
        title="Save output file as",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if save_path:
        try:
            output_file_name = 'output_file.xlsx'
            output_file_path = os.path.join(os.path.expanduser('~'), 'Downloads', output_file_name)
            os.rename(output_file_path, save_path)
            messagebox.showinfo("Success", f"File saved as {save_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save the file: {str(e)}")

# Initialize global variables
input_file_path = ""

# Create a Tkinter root window
root = tk.Tk()
root.title("Audit Sight Template Data Validator")

frame = tk.Frame(root)
frame.pack(padx=10, pady=10)

label = tk.Label(frame, text="Select your Audit Sight Template file:")
label.pack(pady=(0, 10))

select_button = tk.Button(frame, text="Select File", command=select_file)
select_button.pack()

validate_button = tk.Button(frame, text="Validate", state=tk.DISABLED, command=run_validation)
validate_button.pack(pady=(10, 10))

download_button = tk.Button(frame, text="Download Output File", state=tk.DISABLED, command=download_file)
download_button.pack()

log_output = scrolledtext.ScrolledText(frame, width=80, height=20, wrap=tk.WORD)
log_output.pack(pady=(10, 0))

# Run the Tkinter event loop
root.mainloop()
